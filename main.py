import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from datetime import datetime
import copy
import glob
import re
import locale
import os # <--- ADICIONADO

# Configura o locale para português para lidar com nomes de meses
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    print("Aviso: Locale 'pt_BR.UTF-8' não encontrado. Usando o padrão do sistema.")

# === 1. Função principal para processar um arquivo ===
def processar_globo_file(globo_file, wb_destino, template_sheet_name, target_sheet_name):
    print(f"🔄 Processando arquivo: {globo_file}")

    # Checa se a aba já existe para evitar duplicidade
    if target_sheet_name in wb_destino.sheetnames:
        print(f"⚠️ Aviso: A aba '{target_sheet_name}' já existe. Pulando o processamento para este arquivo.")
        return

    # === 1.1 Ler planilha da Globo ===
    try:
        globo_df = pd.read_excel(globo_file)
    except FileNotFoundError:
        print(f"❌ Erro: Arquivo '{globo_file}' não encontrado.")
        return
    except Exception as e:
        print(f"❌ Erro ao ler o arquivo '{globo_file}': {e}")
        return

    abrangencias = ['MAE', 'MAI', 'MA1', 'IMP', 'BAS', 'CDO']
    # Garante que a coluna 'abrangencia' exista antes de filtrar
    if 'abrangencia' not in globo_df.columns:
        print(f"❌ Erro: A coluna 'abrangencia' não foi encontrada no arquivo '{globo_file}'.")
        return
        
    df = globo_df[globo_df['abrangencia'].isin(abrangencias)].copy()

    # Normalizar
    df['mnemonico'] = df['mnemonico'].astype(str).str.upper().str.strip()
    df['horario_inicial'] = df['horario_inicial'].astype(str).str.strip()

    # Corrigir horário
    def formatar_horario(h):
        try:
            return datetime.strptime(str(h)[:5], "%H:%M").strftime("%H:%M")
        except:
            return "-"

    df['horario_fmt'] = df['horario_inicial'].apply(formatar_horario)

    # === 1.2 Pivotar dados ===
    pivot = df.pivot_table(
        index=['mnemonico', 'nome_programa', 'dias_exibicao', 'horario_fmt'],
        columns='abrangencia',
        values='preco_30s',
        aggfunc='first'
    ).reset_index()

    pivot = pivot.rename(columns={
        'mnemonico': 'PROG',
        'nome_programa': 'NOME',
        'dias_exibicao': 'DIA',
        'horario_fmt': 'HORÁRIO'
    })

    pivot['INDICE'] = 0.50
    
    # Adiciona colunas de abrangência que podem não existir no pivot
    for abr in abrangencias:
        if abr not in pivot.columns:
            pivot[abr] = None

    cols = ['PROG', 'NOME', 'DIA', 'HORÁRIO', 'INDICE'] + abrangencias
    pivot = pivot[cols]

    # === 1.3 Ordenação avançada ===
    def get_dia_sort_key(dia):
        dia_ordem = {
            'SEG/SÁB': 0, 'SEG-SAB': 0,
            'SEG/TER/QUA/QUI/SEX': 1, 'SEG-SEX': 1,
            'SEG': 2, 'TER': 3, 'TER/QUI': 4,
            'QUA': 5, 'QUI': 6, 'SEX': 7,
            'SEG/DOM': 8, 'SÁB': 9, 'DOM': 10
        }
        if pd.isna(dia): return 11
        dia_upper = str(dia).strip().upper().replace('SAB', 'SÁB')
        return dia_ordem.get(dia_upper, 11)

    def hora_sort_key(h):
        try:
            return datetime.strptime(str(h), "%H:%M").time()
        except:
            return datetime.strptime("23:59", "%H:%M").time()

    pivot['dia_sort_key'] = pivot['DIA'].apply(get_dia_sort_key)
    pivot['hora_sort_key'] = pivot['HORÁRIO'].apply(hora_sort_key)

    # === 1.4 Separar, ordenar e juntar blocos ===
    pivot['DIA_UPPER'] = pivot['DIA'].astype(str).str.upper().str.replace('SAB', 'SÁB')
    bloco_reaplicacao = pivot[pivot['DIA_UPPER'].isin(['SÁB', 'DOM'])].copy()
    bloco_principal = pivot[~pivot['DIA_UPPER'].isin(['SÁB', 'DOM'])].copy()

    bloco_principal = bloco_principal.sort_values(
        by=['dia_sort_key', 'hora_sort_key'], ignore_index=True
    ).drop(columns=['dia_sort_key', 'hora_sort_key', 'DIA_UPPER'])

    bloco_reaplicacao = bloco_reaplicacao.sort_values(
        by=['dia_sort_key', 'hora_sort_key'], ignore_index=True
    ).drop(columns=['dia_sort_key', 'hora_sort_key', 'DIA_UPPER'])

    linha_reaplicacao = pd.DataFrame([['REAPLICAÇÃO'] + [None] * (len(cols) - 1)], columns=cols)

    final_df = pd.concat([bloco_principal, linha_reaplicacao, bloco_reaplicacao], ignore_index=True)

    # === 1.5 Duplicar aba e preencher ===
    ws_template = wb_destino["TABELA"]
    ws_nova = wb_destino.copy_worksheet(ws_template)
    ws_nova.title = target_sheet_name
    
    # Limpa as linhas de dados antigas da nova aba para evitar lixo
    for row in ws_nova['A3:K100']: # Limpa um intervalo grande
        for cell in row:
            cell.value = None

    start_row = 3
    for r_idx, row_data in final_df.iterrows():
        for c_idx, col_name in enumerate(cols, start=1):
            cell = ws_nova.cell(row=start_row + r_idx, column=c_idx)
            cell.value = row_data[col_name]

    print(f"✅ Aba '{target_sheet_name}' criada e preenchida com sucesso!")


# === 2. Lógica de controle principal ===
def main():
    # --- CAMINHOS AJUSTADOS ---
    PASTA_ENTRADA = "entrada"
    local_file = os.path.join(PASTA_ENTRADA, "PI - REDE MIRANTE.xlsx")
    globo_file_pattern = os.path.join(PASTA_ENTRADA, "Precos Globo_????_??.xlsx")
    
    # Garante que a pasta de entrada exista
    os.makedirs(PASTA_ENTRADA, exist_ok=True)
    
    try:
        wb = load_workbook(local_file)
    except FileNotFoundError:
        print(f"❌ Erro: Arquivo de destino '{local_file}' não encontrado. Verifique se ele está na pasta '{PASTA_ENTRADA}'.")
        return

    if "TABELA" not in wb.sheetnames:
        print("❌ Erro: Aba de template 'TABELA' não encontrada na planilha de destino.")
        return

    globo_files = glob.glob(globo_file_pattern)
    if not globo_files:
        print(f"⚠️ Aviso: Nenhum arquivo correspondente a 'Precos Globo_????_??.xlsx' encontrado na pasta '{PASTA_ENTRADA}'.")
        return

    for globo_file in globo_files:
        match = re.search(r'_(\d{4})_(\d{2})', globo_file)
        if match:
            ano = match.group(1)
            mes = int(match.group(2))
            nome_mes = datetime(int(ano), mes, 1).strftime('%B').upper()
            target_sheet_name = f"TABELA_{nome_mes}_{ano}"
            processar_globo_file(globo_file, wb, "TABELA", target_sheet_name)
        else:
            print(f"❌ Erro: Não foi possível extrair o mês e ano do nome do arquivo: {globo_file}")

    # Salvar o arquivo de destino
    try:
        wb.save(local_file)
        print(f"\n✅ Todos os arquivos processados. '{local_file}' salvo com sucesso!")
    except PermissionError:
        print(f"\n❌ ERRO DE PERMISSÃO: Não foi possível salvar o arquivo '{local_file}'.")
        print("   Por favor, feche o arquivo no Excel e tente novamente.")
    except Exception as e:
        print(f"❌ Erro ao salvar o arquivo '{local_file}': {e}")


if __name__ == "__main__":
    main()