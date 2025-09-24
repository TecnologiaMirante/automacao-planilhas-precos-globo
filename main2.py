import pandas as pd
import openpyxl
import os
import re
from datetime import datetime, timedelta
import locale
import unicodedata
import openpyxl.cell.cell

# --- CONFIGURAÇÕES ---
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    print("Aviso: Locale 'pt_BR.UTF-8' não encontrado.")

PASTA_ENTRADA = "entrada"
PASTA_SAIDA = "saida"
ARQUIVO_DADOS = "PI - REDE MIRANTE.xlsx"
ARQUIVO_MODELO = "Lista de Preços e Patrocínios.xlsx"
ABA_MODELO_NOME = "PREÇOS 30\""

MAPA_ABRANGENCIA = {
    'MA1': 'São Luís', 'MAI': 'Interior', 'MAE': 'Estadual',
    'COD': 'Codó', 'IMP': 'Imperatriz', 'BAS': 'Balsas'
}

# --- FUNÇÕES AUXILIARES ---
def normalizar_string(s):
    if not isinstance(s, str): return ""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper().strip()

def split_tokens(s):
    if not s: return set()
    parts = re.split(r'[\s\/\,;]+', s)
    return {p.strip() for p in parts if p and p.strip()}

def padronizar_dia(dia_str):
    s_norm = normalizar_string(dia_str)
    if not s_norm or s_norm == '-': return '-'
    tokens = split_tokens(s_norm)
    dias_seg_sex = {'SEG', 'TER', 'QUA', 'QUI', 'SEX'}
    dias_seg_sab = {'SEG', 'TER', 'QUA', 'QUI', 'SEX', 'SAB'}
    dias_seg_dom = {'SEG', 'TER', 'QUA', 'QUI', 'SEX', 'SAB', 'DOM'}
    if tokens == dias_seg_dom: return 'SEG-DOM'
    if tokens == dias_seg_sab: return 'SEG-SAB'
    if tokens == dias_seg_sex: return 'SEG-SEX'
    return s_norm.replace(" / ", "/").replace(" , ", ",")

def get_dia_ordem(dia_padronizado_str):
    s = dia_padronizado_str
    ordem = {
        'SEG': 10, 'SEG-SEX': 11, 'SEG-SAB': 12, 'SEG-DOM': 13,
        'TER': 20, 'TER/QUI': 21,
        'QUA': 30, 'QUI': 40, 'SEX': 50, 'SAB': 60, 'DOM': 70, '-': 99,
    }
    return ordem.get(s, 98)

def ajustar_horario_para_ordenacao(horario):
    if not isinstance(horario, datetime): return None
    if horario.hour < 4: return horario + timedelta(hours=24)
    return horario

def get_dia_group(dia_str):
    s = normalizar_string(dia_str)
    tokens = split_tokens(s)
    if tokens == {"SAB"}: return "SAB"
    if tokens == {"DOM"}: return "DOM"
    return "SEG-SEX"

# NOVA FUNÇÃO: Atualiza o rodapé dinamicamente
def atualizar_rodape(sheet, mes, ano):
    hoje = datetime.now().strftime('%d/%m/%Y')
    texto_mes_ano = f'EM {mes.upper()} DE {ano}'
    
    # Procura pelas células do rodapé (começando de uma linha mais baixa)
    for row in sheet.iter_rows(min_row=50, max_row=sheet.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Atualiza o título do rodapé
                if "LISTA DE PREÇOS VÁLIDA" in cell.value:
                    cell.value = f"LISTA DE PREÇOS VÁLIDA PARA COMPRAS REALIZADAS {texto_mes_ano}"
                # Atualiza a data de atualização
                elif "ATUALIZADA EM" in cell.value:
                    cell.value = f"ATUALIZADA EM {hoje}"

# --- FUNÇÃO PRINCIPAL ---
def gerar_relatorios_finais():
    print("--- Iniciando a Geração de Relatórios Finais por Abrangência ---")
    
    caminho_dados = os.path.join(PASTA_ENTRADA, ARQUIVO_DADOS)
    caminho_modelo = os.path.join(PASTA_ENTRADA, ARQUIVO_MODELO)

    if not os.path.exists(caminho_dados) or not os.path.exists(caminho_modelo):
        print(f"ERRO CRÍTICO: Verifique se os arquivos de entrada existem na pasta '{PASTA_ENTRADA}'.")
        return

    xls = pd.ExcelFile(caminho_dados)

    for sheet_name in xls.sheet_names:
        if not sheet_name.startswith("TABELA_"): continue
        match = re.search(r'TABELA_(\w+)_(\d{4})', sheet_name.strip())
        if not match: continue
        
        print(f"\n🔄 Lendo dados da aba: '{sheet_name}'...")
        mes_nome_str, ano_str = match.groups()

        pasta_saida_final = os.path.join(PASTA_SAIDA, "ABRANGENCIAS", ano_str, mes_nome_str.upper())
        os.makedirs(pasta_saida_final, exist_ok=True)

        df_pivotado = pd.read_excel(xls, sheet_name=sheet_name, header=1)
        id_vars = ['PROG', 'NOME', 'DIA', 'HORÁRIO', 'INDICE']
        if not all(col in df_pivotado.columns for col in id_vars):
            print(f"  ❌ ERRO: A aba '{sheet_name}' não contém as colunas esperadas.")
            continue

        df_unpivoted = df_pivotado.melt(id_vars=id_vars, var_name='ABRANGENCIA', value_name='30"').dropna(subset=['30"'])
        df_unpivoted.rename(columns={'PROG': 'SIGLA', 'NOME': 'PROGRAMA', 'HORÁRIO': 'HORARIO'}, inplace=True)
        for col in ['15"', '10"', 'GENERO']: df_unpivoted[col] = ''

        for abr_codigo, abr_nome in MAPA_ABRANGENCIA.items():
            # --- NOVO: VERIFICA SE O ARQUIVO JÁ EXISTE ---
            nome_arquivo_saida = f"{abr_codigo} - Lista de Preços e Patrocínios - {mes_nome_str.capitalize()} {ano_str}.xlsx"
            caminho_saida = os.path.join(pasta_saida_final, nome_arquivo_saida)
            if os.path.exists(caminho_saida):
                print(f"  > Arquivo '{nome_arquivo_saida}' já existe. Pulando...")
                continue
            
            print(f"  > Gerando relatório para: {abr_codigo}...")

            df_final_relatorio = df_unpivoted[df_unpivoted['ABRANGENCIA'] == abr_codigo].copy()
            if df_final_relatorio.empty: continue

            df_final_relatorio['DIA_PADRONIZADO'] = df_final_relatorio['DIA'].apply(padronizar_dia)

            workbook = openpyxl.load_workbook(caminho_modelo)
            sheet = workbook[ABA_MODELO_NOME]

            sheet['A2'] = f'LISTA DE PREÇOS {mes_nome_str.upper()} DE {ano_str}'
            sheet['A3'] = f'{abr_nome.upper()} ({abr_codigo})'

            posicoes_titulos = {}
            primeira_secao_encontrada = False
            for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row):
                val = str(row[0].value).strip().upper() if row[0].value else ""
                if not val: continue
                if "SÁBADO" in val: posicoes_titulos['SAB'] = row[0].row
                elif "DOMINGO" in val: posicoes_titulos['DOM'] = row[0].row
                elif not primeira_secao_encontrada:
                    posicoes_titulos['SEG-SEX'] = row[0].row
                    primeira_secao_encontrada = True
            
            linhas_titulo_ordenadas = sorted(posicoes_titulos.values())
            for i, start_row in enumerate(linhas_titulo_ordenadas):
                clear_start = start_row + 2
                # --- CORREÇÃO: Limita a limpeza para não apagar o rodapé ---
                clear_end = linhas_titulo_ordenadas[i+1] - 2 if i + 1 < len(linhas_titulo_ordenadas) else start_row + 100
                for r in range(clear_start, clear_end + 1):
                    # Adiciona uma verificação para não tentar limpar além do limite da planilha
                    if r > sheet.max_row: break
                    for c in range(1, 9):
                        # Lógica de limpeza (sem alterações)
                        ...

            df_final_relatorio['DIA_GRP'] = df_final_relatorio['DIA'].apply(get_dia_group)
            for tipo_dia, df_grupo in df_final_relatorio.groupby('DIA_GRP'):
                if tipo_dia not in posicoes_titulos: continue
                
                df_grupo['DIA_ORDEM'] = df_grupo['DIA_PADRONIZADO'].apply(get_dia_ordem)
                horarios_convertidos = pd.to_datetime(df_grupo['HORARIO'], format='%H:%M', errors='coerce')
                df_grupo['HORARIO_AJUSTADO'] = horarios_convertidos.apply(ajustar_horario_para_ordenacao)
                df_grupo_ordenado = df_grupo.sort_values(by=['DIA_ORDEM', 'HORARIO_AJUSTADO', 'INDICE'], na_position='last')
                
                linha_atual = posicoes_titulos[tipo_dia] + 2
                for _, row_data in df_grupo_ordenado.iterrows():
                    horario_obj = pd.to_datetime(row_data['HORARIO'], errors='coerce')
                    horario_val = horario_obj.strftime('%H:%M') if pd.notna(horario_obj) else '-'
                    valores = [
                        row_data.get('DIA_PADRONIZADO', ''), horario_val, row_data.get('SIGLA', ''),
                        row_data.get('PROGRAMA', ''), row_data.get('GENERO', ''),
                        row_data.get('30"', 0), row_data.get('15"', 0), row_data.get('10"', 0)
                    ]
                    for col_idx, value in enumerate(valores, 1):
                        cell = sheet.cell(row=linha_atual, column=col_idx)
                        if isinstance(cell, openpyxl.cell.cell.MergedCell): continue
                        cell.value = value
                    linha_atual += 1

            # --- NOVO: ATUALIZA O RODAPÉ ANTES DE SALVAR ---
            atualizar_rodape(sheet, mes_nome_str, ano_str)
            
            workbook.save(caminho_saida)
            print(f"    ✅ Arquivo salvo: {nome_arquivo_saida}")

    print("\n--- Processo Finalizado com Sucesso ---")

if __name__ == '__main__':
    os.makedirs(PASTA_ENTRADA, exist_ok=True)
    os.makedirs(PASTA_SAIDA, exist_ok=True)
    gerar_relatorios_finais()