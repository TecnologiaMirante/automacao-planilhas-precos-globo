import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re
from datetime import datetime, timedelta
import locale
import unicodedata
import openpyxl.cell.cell
import glob

# --- 1. CONFIGURAÇÕES GERAIS ---
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    print("Aviso: Locale 'pt_BR.UTF-8' não encontrado.")

# Pastas e arquivos
PASTA_ENTRADA = "entrada"
PASTA_SAIDA = "saida"

# Arquivos da Etapa 1 (Importação)
ARQUIVO_MESTRE_DADOS = "PI - REDE MIRANTE.xlsx"
PADRAO_ARQUIVOS_GLOBO = "Precos Globo_????_??.xlsx"
TEMPLATE_MESTRE_NOME = "TABELA"

# Arquivos da Etapa 2 (Geração de Relatórios)
ARQUIVO_MODELO_RELATORIO = "Lista de Preços e Patrocínios.xlsx"
ABA_MODELO_RELATORIO_NOME = "PREÇOS 30\""

MAPA_ABRANGENCIA = {
    'MAE': 'Estadual', 'MAI': 'Interior', 'MA1': 'São Luís',
    'IMP': 'Imperatriz', 'BAS': 'Balsas', 'CDO': 'Codó'
}

# --- 2. FUNÇÕES DA ETAPA 1 (Importação e Atualização da Base de Dados) ---

# --- 2. FUNÇÕES DA ETAPA 1 (Importação e Atualização da Base de Dados) ---

def processar_arquivo_globo(globo_file, wb_destino, template_sheet_name, target_sheet_name):
    print(f"  > Processando fonte de dados: {os.path.basename(globo_file)}")

    if target_sheet_name in wb_destino.sheetnames:
        print(f"    - Aviso: A aba '{target_sheet_name}' já existe. Pulando.")
        return

    try:
        globo_df = pd.read_excel(globo_file, dtype={'horario_inicial': str, 'horario_final': str})
    except Exception as e:
        print(f"    - Erro ao ler o arquivo '{globo_file}': {e}")
        return

    # Mapeamento e verificação de colunas essenciais
    abrangencias_map = {'MAE': 'MAE', 'MAI': 'MAI', 'MA1': 'MA1', 'IMP': 'IMP', 'BAS': 'BAS', 'CDO': 'CDO'}
    colunas_essenciais = ['abrangencia', 'mnemonico', 'nome_programa', 'dias_exibicao', 'horario_inicial', 'preco_30s', 'preco_15s', 'preco_10s', 'genero']
    for col in colunas_essenciais:
        if col not in globo_df.columns:
            print(f"    - Erro: A coluna obrigatória '{col}' não foi encontrada em '{globo_file}'.")
            return
    
    # Filtra abrangências e seleciona apenas as colunas que vamos usar
    df = globo_df[globo_df['abrangencia'].isin(abrangencias_map.keys())][colunas_essenciais].copy()
    
    # Normalização
    df['mnemonico'] = df['mnemonico'].astype(str).str.upper().str.strip()
    def formatar_horario(h):
        try: return pd.to_datetime(h, errors='coerce').strftime('%H:%M')
        except:
            try: return datetime.strptime(str(h)[:5], "%H:%M").strftime("%H:%M")
            except: return "-"
    df['horario_fmt'] = df['horario_inicial'].apply(formatar_horario)

    # --- ALTERAÇÃO PRINCIPAL: NÃO PIVOTAR ---
    # Manter os dados em formato "longo" (um registro por linha) é mais simples e robusto.
    # A pivotagem será feita na Etapa 2, apenas quando necessário.
    
    # Renomeia colunas para o padrão final
    df.rename(columns={
        'mnemonico': 'PROG',
        'nome_programa': 'NOME',
        'dias_exibicao': 'DIA',
        'horario_fmt': 'HORÁRIO',
        'abrangencia': 'ABRANGENCIA',
        'genero': 'GENERO',
        'preco_30s': '30"',
        'preco_15s': '15"',
        'preco_10s': '10"'
    }, inplace=True)
    
    # --- Lógica de Ordenação ---
    def get_dia_sort_key(dia):
        dia_ordem = {'SEG/SÁB':0, 'SEG-SAB':0, 'SEG/TER/QUA/QUI/SEX':1, 'SEG-SEX':1, 'SEG':2, 'TER':3, 'TER/QUI':4, 'QUA':5, 'QUI':6, 'SEX':7, 'SEG/DOM':8, 'SÁB':9, 'DOM':10}
        if pd.isna(dia): return 11
        dia_upper = str(dia).strip().upper().replace('SAB', 'SÁB')
        return dia_ordem.get(dia_upper, 11)

    def hora_sort_key(h):
        try: return datetime.strptime(str(h), "%H:%M").time()
        except: return datetime.strptime("23:59", "%H:%M").time()

    df['dia_sort_key'] = df['DIA'].apply(get_dia_sort_key)
    df['hora_sort_key'] = df['HORÁRIO'].apply(hora_sort_key)

    # Ordena o DataFrame final
    final_df = df.sort_values(by=['dia_sort_key', 'hora_sort_key', 'PROG'], ignore_index=True)
    
    # Define a ordem final das colunas
    cols_finais = ['PROG', 'NOME', 'DIA', 'HORÁRIO', 'ABRANGENCIA', 'GENERO', '30"', '15"', '10"']
    final_df = final_df[cols_finais]

    # --- Lógica de escrita no Excel ---
    ws_template = wb_destino[template_sheet_name]
    ws_nova = wb_destino.copy_worksheet(ws_template)
    ws_nova.title = target_sheet_name
    
    # Limpa a área de dados da nova aba (apenas algumas linhas e colunas)
    for row in ws_nova['A2:J150']: 
        for cell in row:
            cell.value = None

    # Escreve o cabeçalho
    for c_idx, col_name in enumerate(cols_finais, start=1):
        ws_nova.cell(row=2, column=c_idx, value=col_name)

    # Escreve os dados
    start_row = 3
    for r_idx, row_data in final_df.fillna('').iterrows():
        for c_idx, col_name in enumerate(cols_finais, start=1):
             ws_nova.cell(row=start_row + r_idx, column=c_idx, value=row_data[col_name])
    
    print(f"    - Sucesso: Aba '{target_sheet_name}' criada com todos os dados.")

def etapa1_atualizar_dados_mestres():
    print("\n--- INICIANDO ETAPA 1: ATUALIZAÇÃO DA BASE DE DADOS MESTRE ---")
    caminho_arquivo_mestre = os.path.join(PASTA_ENTRADA, ARQUIVO_MESTRE_DADOS)
    padrao_arquivos_fonte = os.path.join(PASTA_ENTRADA, PADRAO_ARQUIVOS_GLOBO)
    
    try:
        wb = load_workbook(caminho_arquivo_mestre)
    except FileNotFoundError:
        print(f"❌ Erro Crítico: Arquivo mestre '{caminho_arquivo_mestre}' não encontrado.")
        return False

    if TEMPLATE_MESTRE_NOME not in wb.sheetnames:
        print(f"❌ Erro Crítico: Aba de template '{TEMPLATE_MESTRE_NOME}' não encontrada no arquivo mestre.")
        return False

    arquivos_globo = glob.glob(padrao_arquivos_fonte)
    if not arquivos_globo:
        print("- Aviso: Nenhum arquivo de dados brutos ('Precos Globo_...') encontrado para processar.")
        return True # Não é um erro, apenas não há nada a fazer

    ano_max, mes_max = 0, 0
    for globo_file in arquivos_globo:
        match = re.search(r'_(\d{4})_(\d{2})', globo_file)
        if match:
            ano, mes_num = match.groups()
            # Encontra o ano/mês mais recente entre os arquivos de entrada
            if int(ano) > ano_max or (int(ano) == ano_max and int(mes_num) > mes_max):
                ano_max, mes_max = int(ano), int(mes_num)
            
            nome_mes = datetime(int(ano), int(mes_num), 1).strftime('%B').upper()
            target_sheet_name = f"TABELA_{nome_mes}_{ano}"
            processar_arquivo_globo(globo_file, wb, TEMPLATE_MESTRE_NOME, target_sheet_name)

    # --- Lógica de salvamento na pasta de saída ---
    # 1. Usa o mês/ano mais recente encontrado nos arquivos de entrada
    if ano_max == 0: # Caso nenhum arquivo válido tenha sido processado
        print("- Nenhum arquivo de dados válido processado. O arquivo mestre não será salvo.")
        return True # Retorna sucesso, pois não houve erro, apenas nada a fazer.

    nome_mes_recente = datetime(ano_max, mes_max, 1).strftime('%B').capitalize()
    
    # 2. Montar o nome e o caminho do novo arquivo de saída na estrutura de pastas correta
    pasta_saida_ano = os.path.join(PASTA_SAIDA, "PI", str(ano_max))
    os.makedirs(pasta_saida_ano, exist_ok=True)
    nome_arquivo_saida = f"PI - REDE MIRANTE - {nome_mes_recente} {ano_max}.xlsx"
    caminho_saida = os.path.join(pasta_saida_ano, nome_arquivo_saida)

    try:
        wb.save(caminho_saida)
        print(f"✅ ETAPA 1 CONCLUÍDA: Novo arquivo mestre salvo em '{caminho_saida}'")
        return caminho_saida # Retorna o caminho do novo arquivo em caso de sucesso
    except PermissionError:
        print(f"❌ ERRO DE PERMISSÃO: Não foi possível salvar '{caminho_saida}'. Verifique se o arquivo já existe e está aberto.")
        return None # Retorna None em caso de falha
    except Exception as e:
        print(f"❌ Erro ao salvar o novo arquivo mestre: {e}")
        return None # Retorna None em caso de falha


# --- 3. FUNÇÕES DA ETAPA 2 (Geração de Relatórios Finais) ---

def normalizar_string(s):
    if not isinstance(s, str): return ""
    s = unicodedata.normalize("NFD", s); s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.upper().strip()

def split_tokens(s):
    if not s: return set()
    parts = re.split(r'[\s\/\,;]+', s)
    return {p.strip() for p in parts if p and p.strip()}

def padronizar_dia(dia_str):
    s_norm = normalizar_string(dia_str)
    if not s_norm or s_norm == '-': return '-'
    tokens = split_tokens(s_norm)
    dias_seg_sex = {'SEG', 'TER', 'QUA', 'QUI', 'SEX'}; dias_seg_sab = dias_seg_sex | {'SAB'}; dias_seg_dom = dias_seg_sab | {'DOM'}
    if tokens == dias_seg_dom: return 'SEG-DOM'
    if tokens == dias_seg_sab: return 'SEG-SAB'
    if tokens == dias_seg_sex: return 'SEG-SEX'
    return s_norm.replace(" / ", "/").replace(" , ", ",")

def get_dia_ordem(dia_padronizado_str):
    s = dia_padronizado_str
    ordem = {'SEG': 10, 'SEG-SEX': 11, 'SEG-SAB': 12, 'SEG-DOM': 13, 'TER': 20, 'TER/QUI': 21, 'QUA': 30, 'QUI': 40, 'SEX': 50, 'SAB': 60, 'DOM': 70, '-': 99}
    return ordem.get(s, 98)

def ajustar_horario_para_ordenacao(horario):
    if not isinstance(horario, datetime): return None
    if horario.hour < 4: return horario + timedelta(hours=24)
    return horario

def get_dia_group(dia_str):
    tokens = split_tokens(normalizar_string(dia_str))
    if tokens == {"SAB"}: return "SAB"
    if tokens == {"DOM"}: return "DOM"
    return "SEG-SEX"

def atualizar_rodape(sheet, mes, ano):
    hoje = datetime.now().strftime('%d/%m/%Y')
    texto_mes_ano = f'EM {mes.upper()} DE {ano}'
    for row in sheet.iter_rows(min_row=50, max_row=sheet.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if "LISTA DE PREÇOS VÁLIDA" in cell.value: cell.value = f"LISTA DE PREÇOS VÁLIDA PARA COMPRAS REALIZADAS {texto_mes_ano}"
                elif "ATUALIZADA EM" in cell.value: cell.value = f"ATUALIZADA EM {hoje}"

def etapa2_gerar_relatorios_finais(caminho_dados_atualizado):
    print("\n--- INICIANDO ETAPA 2: GERAÇÃO DOS RELATÓRIOS FINAIS ---")
    caminho_modelo = os.path.join(PASTA_ENTRADA, ARQUIVO_MODELO_RELATORIO)

    if not os.path.exists(caminho_dados_atualizado) or not os.path.exists(caminho_modelo):
        print(f"❌ Erro Crítico: Verifique se os arquivos de entrada existem.")
        return

    xls = pd.ExcelFile(caminho_dados_atualizado)

    for sheet_name in xls.sheet_names:
        if not sheet_name.startswith("TABELA_"): continue
        match = re.search(r'TABELA_(\w+)_(\d{4})', sheet_name.strip())
        if not match: continue
        
        print(f"\n> Processando relatórios para a aba: '{sheet_name}'...")
        mes_nome_str, ano_str = match.groups()

        pasta_saida_final = os.path.join(PASTA_SAIDA, "ABRANGENCIAS", ano_str, mes_nome_str.upper())
        os.makedirs(pasta_saida_final, exist_ok=True)

        df_completo = pd.read_excel(xls, sheet_name=sheet_name, header=1)
        df_completo.rename(columns={'PROG': 'SIGLA', 'NOME': 'PROGRAMA', 'HORÁRIO': 'HORARIO'}, inplace=True)

        for abr_codigo, abr_nome in MAPA_ABRANGENCIA.items():
            nome_arquivo_saida = f"{abr_codigo} - Lista de Preços e Patrocínios - {mes_nome_str.capitalize()} {ano_str}.xlsx"
            caminho_saida = os.path.join(pasta_saida_final, nome_arquivo_saida)
            if os.path.exists(caminho_saida):
                print(f"  - Relatório '{nome_arquivo_saida}' já existe. Pulando...")
                continue
            
            print(f"  - Gerando relatório para: {abr_codigo}...")
            
            df_filtrado_abrangencia = df_completo[df_completo['ABRANGENCIA'] == abr_codigo].copy()
            if df_filtrado_abrangencia.empty: continue

            # --- CORREÇÃO PRINCIPAL: REMOVER DUPLICATAS ---
            # Agrupa por programa e pega apenas a primeira entrada.
            # Isso garante que cada programa apareça apenas uma vez.
            df_final_relatorio = df_filtrado_abrangencia.drop_duplicates(
                subset=['SIGLA', 'PROGRAMA', 'DIA', 'HORARIO'],
                keep='first'
            ).copy()
            # --------------------------------------------------

            df_final_relatorio['DIA_PADRONIZADO'] = df_final_relatorio['DIA'].apply(padronizar_dia)
            workbook = openpyxl.load_workbook(caminho_modelo)
            sheet = workbook[ABA_MODELO_RELATORIO_NOME]
            sheet['A2'] = f'LISTA DE PREÇOS {mes_nome_str.upper()} DE {ano_str}'
            sheet['A3'] = f'{abr_nome.upper()} ({abr_codigo})'

            # (O restante da função continua exatamente igual)
            posicoes_titulos = {}
            primeira_secao_encontrada = False
            for row in sheet.iter_rows(min_row=7, max_row=sheet.max_row):
                val = str(row[0].value).strip().upper() if row[0].value else ""
                if not val: continue
                if "SÁBADO" in val: posicoes_titulos['SAB'] = row[0].row
                elif "DOMINGO" in val: posicoes_titulos['DOM'] = row[0].row
                elif not primeira_secao_encontrada:
                    posicoes_titulos['SEG-SEX'] = row[0].row
                    primeira_secao_encontrada = True
            
            linhas_titulo_ordenadas = sorted(posicoes_titulos.values())
            for i, start_row in enumerate(linhas_titulo_ordenadas):
                clear_start = start_row + 2
                clear_end = linhas_titulo_ordenadas[i+1] - 2 if i + 1 < len(linhas_titulo_ordenadas) else start_row + 150
                for r in range(clear_start, clear_end + 1):
                    if r > sheet.max_row: break
                    for c in range(1, 9):
                        cell = sheet.cell(row=r, column=c)
                        if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                            cell.value = None

            df_final_relatorio['DIA_GRP'] = df_final_relatorio['DIA'].apply(get_dia_group)
            for tipo_dia, df_grupo in df_final_relatorio.groupby('DIA_GRP'):
                if tipo_dia not in posicoes_titulos: continue
                
                df_grupo['DIA_ORDEM'] = df_grupo['DIA_PADRONIZADO'].apply(get_dia_ordem)
                horarios_convertidos = pd.to_datetime(df_grupo['HORARIO'], format='%H:%M', errors='coerce')
                df_grupo['HORARIO_AJUSTADO'] = horarios_convertidos.apply(ajustar_horario_para_ordenacao)
                
                df_grupo_ordenado = df_grupo.sort_values(by=['DIA_ORDEM', 'HORARIO_AJUSTADO', 'SIGLA'])
                
                linha_atual = posicoes_titulos[tipo_dia] + 2
                for _, row_data in df_grupo_ordenado.iterrows():
                    horario_obj = pd.to_datetime(row_data['HORARIO'], errors='coerce')
                    horario_val = horario_obj.strftime('%H:%M') if pd.notna(horario_obj) else '-'
                    valores = [
                        row_data.get('DIA_PADRONIZADO', ''),
                        horario_val,
                        row_data.get('SIGLA', ''),
                        row_data.get('PROGRAMA', ''),
                        row_data.get('GENERO', ''),
                        row_data.get('30"', 0),
                        row_data.get('15"', 0),
                        row_data.get('10"', 0)
                    ]
                    for col_idx, value in enumerate(valores, 1):
                        cell = sheet.cell(row=linha_atual, column=col_idx)
                        if isinstance(cell, openpyxl.cell.cell.MergedCell): continue
                        cell.value = value
                    linha_atual += 1
            
            atualizar_rodape(sheet, mes_nome_str, ano_str)
            workbook.save(caminho_saida)
            print(f"    - Sucesso: Relatório '{nome_arquivo_saida}' salvo.")
    print("✅ ETAPA 2 CONCLUÍDA: Geração de relatórios finalizada.")

# --- 4. LÓGICA DE CONTROLE PRINCIPAL ---
def main():
    os.makedirs(PASTA_ENTRADA, exist_ok=True)
    os.makedirs(PASTA_SAIDA, exist_ok=True)

    caminho_arquivo_gerado = etapa1_atualizar_dados_mestres()
    
    if caminho_arquivo_gerado:
        etapa2_gerar_relatorios_finais(caminho_arquivo_gerado)

    print("\n--- Processo Geral Finalizado ---")

if __name__ == "__main__":
    main()